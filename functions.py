import pyautogui as main
import pyperclip as pc
import time
from openpyxl import load_workbook
import subprocess
from points import data, AN, data_nombres


Excel = "GESTION.xlsx"

# Conexión con Openpyxl
Wb = load_workbook(Excel, data_only=True)
sheet = Wb["Hoja1"]

#wait time
main.PAUSE = 0.5
#coordenadas
POSITION_PORCENT =  634,101
PORCENT_VALUE = "45"

#Inicio, busqueda de archivo por medio de comandos y función (empresa) del archivo points.


def inicio():
	main.hotkey("winleft", "d")
	r = sheet['C1'].value
	resultado = r.split()[0]
	Open_pdf = data.get(resultado)
	subprocess.Popen([Open_pdf], shell=True)
	time.sleep(5)


def eliminar_fila():
	sheet.delete_rows(2, sheet.max_row-1)
	Wb.save(Excel) 


#preparing size pdf
def precopy():
	puntero(POSITION_PORCENT)
	main.write(PORCENT_VALUE)
	main.press('enter', presses=1)
	time.sleep(3)

#Clicks
def puntero(pos, click=1):
    main.moveTo(pos)
    main.click(clicks=click)

#data extractor for pyautogui
def function_copy(a):	
	pc.copy(a)
	main.hotkey('ctrl', 'v')

# eliminar fila




