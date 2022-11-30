from ctypes import Array
from openpyxl import load_workbook
import datetime
import pyautogui as main
from points import data_firmas, AN, data_nombres, data_names_

Excel = "GESTION.xlsx"

# Conexión con Openpyxl
Wb = load_workbook(Excel, data_only=True)
sheet = Wb["Hoja1"]


#Extracción de datos por medio de iterador primerafila...

initial_data = [""] * 2 # Se genera doble espacio vacio inicialmente...


def total(a):
    fecha_nac = sheet[a].value
    fecha_dia = fecha_nac.strftime("%m")
    fecha_mes = fecha_nac.strftime("%d")
    totality = fecha_dia , fecha_mes
    totality = [*totality]
    return totality


valor1 = total('D1')
valor2 = total('E1')


def fila_excel():
    delete = [0,2,9] # Datos descartados
    for value in sheet.iter_rows(min_row = 1, max_row = 1,values_only=True):
            elementos = list(value)
            for i in sorted(delete, reverse = True):
                del elementos[i]
    
    elementos[1] = elementos[1].strftime("%Y")# cambio de formato de date time a string
    elementos[2] = elementos[2].strftime("%Y")
    elementos = elementos + valor1 + valor2
    elementos = list(filter(None, elementos))
    elementos = initial_data + elementos + data_firmas + data_names_
    
    
    return elementos
